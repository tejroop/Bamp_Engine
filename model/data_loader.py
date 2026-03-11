#!/usr/bin/env python3
"""
BAMP Dashboard Data Loader
Comprehensive script to read and aggregate real CSV/XLSX data for the BAMP dashboard.
Processes order data, competitor pricing, traffic, and SKU mappings.
"""

import csv
import json
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

# Try to import openpyxl for Excel files, but we'll handle it gracefully
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not available. Excel files will be skipped.")

# Configuration
DATA_DIR = "/sessions/gallant-vigilant-knuth/mnt/2025-11-11 ML Lab MCT 7/"
OUTPUT_DIR = "/sessions/gallant-vigilant-knuth/bamp-engine/backend/data/"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "real_data.json")

# SKU prefix mappings for product categories
SKU_PREFIX_MAPPING = {
    'EMA': 'Mattresses',
    'EMH': 'Mattresses',
    'EMS': 'Mattresses',
    'EMO': 'Mattresses',
    'EMI': 'Mattresses',
    'EPW': 'Pillows',
    'EPC': 'Pillows',
    'EPP': 'Pillows',
    'EDB': 'Bedding',
    'EDD': 'Bedding',
    'EDL': 'Bedding',
    'EDS': 'Bedding',
    'EDA': 'Bedding',
    'EFB': 'Frames',
    'EFS': 'Frames',
    'EFA': 'Frames',
    'EBA': 'Accessories',
    'EBL': 'Accessories',
    'EBC': 'Accessories',
    'EBD': 'Accessories',
}

def parse_date(date_str, formats=None):
    """Parse date string with multiple possible formats."""
    if not date_str or date_str.strip() == '':
        return None
    
    if formats is None:
        formats = ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%m-%d-%Y']
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    
    return None

def read_csv_safe(filepath, encoding='utf-8-sig'):
    """Safely read CSV file with proper encoding handling."""
    try:
        with open(filepath, 'r', encoding=encoding, errors='replace') as f:
            reader = csv.DictReader(f)
            rows = [row for row in reader]
        return rows
    except Exception as e:
        print(f"ERROR reading {filepath}: {e}")
        return []

def get_sku_category(sku):
    """Get product category from SKU prefix."""
    if not sku:
        return 'Unknown'
    
    for prefix, category in SKU_PREFIX_MAPPING.items():
        if sku.upper().startswith(prefix):
            return category
    
    return 'Other'

def process_hk_tw_orders(filepath, market):
    """Process HK and TW order files (202301-202503)."""
    print(f"  Processing {market} order file: {filepath}")
    
    rows = read_csv_safe(filepath)
    if not rows:
        return {
            'monthly_demand': [],
            'attachment_rates': [],
            'products': [],
            'raw_count': 0
        }
    
    monthly_data = defaultdict(lambda: {
        'mattress_qty': 0,
        'accessory_qty': 0,
        'other_qty': 0,
        'total_qty': 0,
        'total_revenue': 0,
        'total_margin': 0,
        'orders': set(),
        'mattress_orders': set(),
        'crosssell_orders': set(),
        'products': defaultdict(lambda: {'qty': 0, 'revenue': 0, 'margin': 0})
    })
    
    for row in rows:
        try:
            date_str = row.get('Order Date', '').strip()
            if not date_str:
                continue
            
            order_date = parse_date(date_str)
            if not order_date:
                continue
            
            month_key = order_date.strftime('%Y-%m')
            product_group = row.get('Product Group L1', '').strip()
            sku = row.get('Sku', '').strip()
            order_id = row.get('Order Id', '').strip()
            
            # Parse quantities and revenue
            try:
                qty = float(row.get('Qty Ordered (Products)', 0) or 0)
                revenue = float(row.get('Net Revenue', 0) or 0)
                margin = float(row.get('Gross Margin', 0) or 0)
            except (ValueError, TypeError):
                continue
            
            if qty == 0 or revenue == 0:
                continue
            
            monthly_data[month_key]['total_qty'] += qty
            monthly_data[month_key]['total_revenue'] += revenue
            monthly_data[month_key]['total_margin'] += margin
            
            if order_id:
                monthly_data[month_key]['orders'].add(order_id)
            
            # Categorize by product group
            if 'Mattress' in product_group or 'mattress' in product_group.lower():
                monthly_data[month_key]['mattress_qty'] += qty
                if order_id:
                    monthly_data[month_key]['mattress_orders'].add(order_id)
            elif 'Accessory' in product_group or 'accessory' in product_group.lower():
                monthly_data[month_key]['accessory_qty'] += qty
            else:
                monthly_data[month_key]['other_qty'] += qty
            
            # Track products
            if sku:
                monthly_data[month_key]['products'][sku]['qty'] += qty
                monthly_data[month_key]['products'][sku]['revenue'] += revenue
                monthly_data[month_key]['products'][sku]['margin'] += margin
        
        except Exception as e:
            continue
    
    # Build output structures
    monthly_demand = []
    for month_key in sorted(monthly_data.keys()):
        data = monthly_data[month_key]
        monthly_demand.append({
            'month': month_key,
            'mattress_qty': data['mattress_qty'],
            'accessory_qty': data['accessory_qty'],
            'other_qty': data['other_qty'],
            'total_qty': data['total_qty'],
            'total_revenue': round(data['total_revenue'], 2),
            'total_margin': round(data['total_margin'], 2),
            'order_count': len(data['orders'])
        })
    
    # Calculate attachment rates
    attachment_rates = []
    for month_key in sorted(monthly_data.keys()):
        data = monthly_data[month_key]
        mattress_orders = len(data['mattress_orders'])
        
        if mattress_orders > 0:
            crosssell_count = len(data['orders']) - mattress_orders
            rate = (crosssell_count / mattress_orders) if mattress_orders > 0 else 0
        else:
            crosssell_count = 0
            rate = 0
        
        attachment_rates.append({
            'month': month_key,
            'rate': round(rate, 4),
            'mattress_orders': mattress_orders,
            'crosssell_orders': crosssell_count
        })
    
    # Build product list
    products = []
    for month_key in monthly_data.keys():
        for sku, stats in monthly_data[month_key]['products'].items():
            products.append({
                'sku': sku,
                'month': month_key,
                'qty': stats['qty'],
                'revenue': round(stats['revenue'], 2),
                'margin': round(stats['margin'], 2),
                'category': get_sku_category(sku)
            })
    
    return {
        'monthly_demand': monthly_demand,
        'attachment_rates': attachment_rates,
        'products': products,
        'raw_count': len(rows)
    }

def process_recent_orders(filepath):
    """Process recent TW&HK order data (202504-202512)."""
    print(f"  Processing recent order file: {filepath}")
    
    rows = read_csv_safe(filepath)
    if not rows:
        return {
            'monthly_demand': [],
            'products': [],
            'raw_count': 0
        }
    
    monthly_data = defaultdict(lambda: {
        'total_qty': 0,
        'total_revenue': 0,
        'total_margin': 0,
        'orders': set(),
        'products': defaultdict(lambda: {'qty': 0, 'revenue': 0, 'margin': 0})
    })
    
    for row in rows:
        try:
            date_str = row.get('Order Created At Date', '').strip()
            if not date_str:
                continue
            
            order_date = parse_date(date_str)
            if not order_date:
                continue
            
            month_key = order_date.strftime('%Y-%m')
            sku = row.get('Product Sku', '').strip()
            order_id = row.get('Order Ref', '').strip()
            
            # Parse quantities and revenue
            try:
                qty = float(row.get('Qty Net', 0) or 0)
                revenue = float(row.get('Net Revenue Eur', 0) or 0)
                margin = float(row.get('Gross Margin Eur', 0) or 0)
            except (ValueError, TypeError):
                continue
            
            if qty == 0 or revenue == 0:
                continue
            
            monthly_data[month_key]['total_qty'] += qty
            monthly_data[month_key]['total_revenue'] += revenue
            monthly_data[month_key]['total_margin'] += margin
            
            if order_id:
                monthly_data[month_key]['orders'].add(order_id)
            
            # Track products
            if sku:
                monthly_data[month_key]['products'][sku]['qty'] += qty
                monthly_data[month_key]['products'][sku]['revenue'] += revenue
                monthly_data[month_key]['products'][sku]['margin'] += margin
        
        except Exception as e:
            continue
    
    # Build output structures
    monthly_demand = []
    for month_key in sorted(monthly_data.keys()):
        data = monthly_data[month_key]
        monthly_demand.append({
            'month': month_key,
            'total_qty': data['total_qty'],
            'total_revenue': round(data['total_revenue'], 2),
            'total_margin': round(data['total_margin'], 2),
            'order_count': len(data['orders'])
        })
    
    # Build product list
    products = []
    for month_key in monthly_data.keys():
        for sku, stats in monthly_data[month_key]['products'].items():
            products.append({
                'sku': sku,
                'month': month_key,
                'qty': stats['qty'],
                'revenue': round(stats['revenue'], 2),
                'margin': round(stats['margin'], 2),
                'category': get_sku_category(sku)
            })
    
    return {
        'monthly_demand': monthly_demand,
        'products': products,
        'raw_count': len(rows)
    }

def process_competitor_prices(filepath):
    """Process competitor price data."""
    print(f"  Processing competitor prices: {filepath}")
    
    rows = read_csv_safe(filepath)
    if not rows:
        return {}
    
    market_data = defaultdict(lambda: defaultdict(lambda: {
        'brands': set(),
        'prices': [],
        'count': 0
    }))
    
    for row in rows:
        try:
            country = row.get('country', '').strip()
            week = row.get('week', '').strip()
            year = row.get('year', '').strip()
            brand = row.get('brand', '').strip()
            
            if not all([country, week, year]):
                continue
            
            try:
                discounted_price = float(row.get('discounted_price', 0) or 0)
                original_price = float(row.get('original_price', 0) or 0)
            except (ValueError, TypeError):
                continue
            
            if discounted_price <= 0:
                continue
            
            key = (int(week), int(year))
            market_data[country][key]['brands'].add(brand)
            market_data[country][key]['prices'].append(discounted_price)
            market_data[country][key]['count'] += 1
        
        except Exception as e:
            continue
    
    # Build output
    result = {}
    for country, weeks in market_data.items():
        result[country] = []
        for (week, year), data in sorted(weeks.items()):
            prices = data['prices']
            if prices:
                result[country].append({
                    'week': week,
                    'year': year,
                    'brands': list(data['brands']),
                    'avg_discounted': round(sum(prices) / len(prices), 2),
                    'min_price': round(min(prices), 2),
                    'max_price': round(max(prices), 2),
                    'competitor_count': len(data['brands'])
                })
    
    return result

def process_discount_codes(filepath):
    """Process discount code data."""
    print(f"  Processing discount codes: {filepath}")
    
    rows = read_csv_safe(filepath)
    if not rows:
        return {}
    
    category_data = defaultdict(lambda: {
        'discounts': [],
        'count': 0
    })
    
    for row in rows:
        try:
            product_group = row.get('product_group_l1', '').strip()
            sku = row.get('sku', row.get('product_sku', '')).strip()
            
            if not product_group:
                product_group = get_sku_category(sku)
            
            try:
                original_price = float(row.get('original_unit_price', 0) or 0)
                discount_amount = float(row.get('unit_discount_amount_euro', 0) or 0)
            except (ValueError, TypeError):
                continue
            
            if original_price <= 0:
                continue
            
            discount_pct = (discount_amount / original_price) * 100
            category_data[product_group]['discounts'].append(discount_pct)
            category_data[product_group]['count'] += 1
        
        except Exception as e:
            continue
    
    # Build output
    result = {}
    if category_data:
        all_discounts = []
        for category, data in category_data.items():
            if data['discounts']:
                avg = sum(data['discounts']) / len(data['discounts'])
                result[category] = {
                    'avg_discount_pct': round(avg, 2),
                    'count': data['count']
                }
                all_discounts.extend(data['discounts'])
        
        if all_discounts:
            result['overall'] = {
                'avg_discount_pct': round(sum(all_discounts) / len(all_discounts), 2),
                'count': len(all_discounts)
            }
    
    return result

def process_traffic(filepath, market):
    """Process traffic data."""
    print(f"  Processing {market} traffic: {filepath}")
    
    rows = read_csv_safe(filepath)
    if not rows:
        return []
    
    monthly_traffic = defaultdict(lambda: {
        'visits': 0,
        'orders': 0,
        'count': 0
    })
    
    for row in rows:
        try:
            date_str = row.get('Order Date', '').strip()
            if not date_str:
                continue
            
            order_date = parse_date(date_str)
            if not order_date:
                continue
            
            month_key = order_date.strftime('%Y-%m')
            
            try:
                visits = float(row.get('Site Visits', 0) or 0)
                orders = float(row.get('Orders', 0) or 0)
            except (ValueError, TypeError):
                continue
            
            if visits > 0:
                monthly_traffic[month_key]['visits'] += visits
                monthly_traffic[month_key]['orders'] += orders
                monthly_traffic[month_key]['count'] += 1
        
        except Exception as e:
            continue
    
    # Build output
    traffic_list = []
    for month_key in sorted(monthly_traffic.keys()):
        data = monthly_traffic[month_key]
        visits = data['visits']
        orders = data['orders']
        conversion = (orders / visits) * 100 if visits > 0 else 0
        
        traffic_list.append({
            'month': month_key,
            'visits': int(visits),
            'orders': int(orders),
            'conversion_rate': round(conversion, 2)
        })
    
    return traffic_list

def process_sku_labelling(filepath):
    """Process SKU labelling from Excel file."""
    print(f"  Processing SKU labelling: {filepath}")
    
    if not OPENPYXL_AVAILABLE:
        print("    Skipping - openpyxl not available")
        return {}
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        sku_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3:
                sku3 = row[0]
                sku5 = row[1]
                product_name = row[2]
                
                if sku5:
                    prefix = str(sku5)[:3].upper()
                    if prefix not in sku_map and product_name:
                        sku_map[prefix] = product_name
        
        return sku_map
    except Exception as e:
        print(f"    Error reading Excel: {e}")
        return {}

def main():
    """Main execution function."""
    print("=" * 70)
    print("BAMP Dashboard Data Loader")
    print("=" * 70)
    print(f"\nData Directory: {DATA_DIR}")
    print(f"Output Directory: {OUTPUT_DIR}\n")
    
    # Initialize data structure
    data = {
        'generated_at': datetime.now().isoformat(),
        'data_sources': [],
        'markets': {
            'HK': {
                'monthly_demand': [],
                'attachment_rates': [],
                'products': [],
                'traffic': []
            },
            'TW': {
                'monthly_demand': [],
                'attachment_rates': [],
                'products': [],
                'traffic': []
            }
        },
        'competitor_data': {},
        'sku_mapping': {},
        'discount_analysis': {},
        'summary': {}
    }
    
    # Process HK and TW orders (2023-2025)
    hk_file = os.path.join(DATA_DIR, "HK 202301-202503.csv")
    if os.path.exists(hk_file):
        print("\n[1] Processing HK Orders (2023-2025)")
        hk_orders = process_hk_tw_orders(hk_file, "HK")
        data['markets']['HK']['monthly_demand'] = hk_orders['monthly_demand']
        data['markets']['HK']['attachment_rates'] = hk_orders['attachment_rates']
        data['markets']['HK']['products'] = hk_orders['products']
        data['data_sources'].append("HK 202301-202503.csv")
        print(f"    Processed {hk_orders['raw_count']} rows")
    
    tw_file = os.path.join(DATA_DIR, "TW 202301-202503.csv")
    if os.path.exists(tw_file):
        print("\n[2] Processing TW Orders (2023-2025)")
        tw_orders = process_hk_tw_orders(tw_file, "TW")
        data['markets']['TW']['monthly_demand'] = tw_orders['monthly_demand']
        data['markets']['TW']['attachment_rates'] = tw_orders['attachment_rates']
        data['markets']['TW']['products'] = tw_orders['products']
        data['data_sources'].append("TW 202301-202503.csv")
        print(f"    Processed {tw_orders['raw_count']} rows")
    
    # Process recent orders
    recent_file = os.path.join(DATA_DIR, "TW&HK 202504-202512 Order data.csv")
    if os.path.exists(recent_file):
        print("\n[3] Processing Recent TW&HK Orders (2025-2026)")
        recent_orders = process_recent_orders(recent_file)
        # Merge recent data (append to existing or create separate section)
        data['data_sources'].append("TW&HK 202504-202512 Order data.csv")
        print(f"    Processed {recent_orders['raw_count']} rows")
    
    # Process traffic
    hk_traffic_file = os.path.join(DATA_DIR, "HK-Traffic.csv")
    if os.path.exists(hk_traffic_file):
        print("\n[4] Processing HK Traffic")
        hk_traffic = process_traffic(hk_traffic_file, "HK")
        data['markets']['HK']['traffic'] = hk_traffic
        data['data_sources'].append("HK-Traffic.csv")
        print(f"    Processed {len(hk_traffic)} months of data")
    
    tw_traffic_file = os.path.join(DATA_DIR, "TW-Traffic.csv")
    if os.path.exists(tw_traffic_file):
        print("\n[5] Processing TW Traffic")
        tw_traffic = process_traffic(tw_traffic_file, "TW")
        data['markets']['TW']['traffic'] = tw_traffic
        data['data_sources'].append("TW-Traffic.csv")
        print(f"    Processed {len(tw_traffic)} months of data")
    
    # Process competitor prices
    competitor_file = os.path.join(DATA_DIR, "Competitor price.csv")
    if os.path.exists(competitor_file):
        print("\n[6] Processing Competitor Prices")
        competitor_data = process_competitor_prices(competitor_file)
        data['competitor_data'] = competitor_data
        data['data_sources'].append("Competitor price.csv")
        for market, records in competitor_data.items():
            print(f"    {market}: {len(records)} weeks of data")
    
    # Process discount codes
    discount_file = os.path.join(DATA_DIR, "TW-2023-2024 discount code.csv")
    if os.path.exists(discount_file):
        print("\n[7] Processing Discount Codes")
        discount_data = process_discount_codes(discount_file)
        data['discount_analysis']['TW'] = discount_data
        data['data_sources'].append("TW-2023-2024 discount code.csv")
        print(f"    Processed {len(discount_data)} categories")
    
    # Process SKU labelling
    sku_file = os.path.join(DATA_DIR, "SKU labelling.xlsx")
    if os.path.exists(sku_file):
        print("\n[8] Processing SKU Labelling")
        sku_map = process_sku_labelling(sku_file)
        data['sku_mapping'] = sku_map
        data['data_sources'].append("SKU labelling.xlsx")
        print(f"    Mapped {len(sku_map)} SKU prefixes")
    
    # Calculate summary statistics
    print("\n[9] Calculating Summary Statistics")
    for market in ['HK', 'TW']:
        market_data = data['markets'][market]
        demand = market_data['monthly_demand']
        attachment = market_data['attachment_rates']
        
        if demand:
            total_orders = sum(d['order_count'] for d in demand)
            total_revenue = sum(d['total_revenue'] for d in demand)
            avg_attachment = (sum(a['rate'] for a in attachment) / len(attachment)) if attachment else 0
            
            date_range = f"{demand[0]['month']} to {demand[-1]['month']}"
            
            data['summary'][market] = {
                'total_orders': total_orders,
                'total_revenue': round(total_revenue, 2),
                'avg_attachment_rate': round(avg_attachment, 4),
                'date_range': date_range,
                'months_processed': len(demand)
            }
            
            print(f"  {market}: {total_orders} orders, {total_revenue:.2f} EUR, {avg_attachment:.2%} attachment")
    
    # Write output
    print("\n[10] Writing Output JSON")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"  Output written to: {OUTPUT_FILE}")
    print(f"  File size: {os.path.getsize(OUTPUT_FILE)} bytes")
    
    print("\n" + "=" * 70)
    print("PROCESSING COMPLETE")
    print("=" * 70)
    print(f"\nData Sources Processed: {len(data['data_sources'])}")
    for source in data['data_sources']:
        print(f"  - {source}")
    
    print("\nSummary Statistics:")
    for market, summary in data['summary'].items():
        print(f"\n  {market}:")
        for key, value in summary.items():
            print(f"    {key}: {value}")
    
    return data

if __name__ == '__main__':
    main()
